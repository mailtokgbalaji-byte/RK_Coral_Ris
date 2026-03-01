import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

# --- 1. APP CONFIGURATION ---
st.set_page_config(
    page_title="RK Event Reporter Pro",
    page_icon="🛠️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for a professional look
st.markdown("""
    <style>
    .main { background-color: #f5f7f9; }
    .stButton>button { 
        width: 100%; 
        border-radius: 5px; 
        height: 3em; 
        background-color: #ff4b4b; 
        color: white; 
        font-weight: bold;
    }
    .stButton>button:hover {
        background-color: #e04444;
    }
    .stDataFrame { border: 1px solid #e6e9ef; border-radius: 5px; }
    </style>
    """, unsafe_allow_html=True)

# --- 2. HELPER FUNCTIONS ---
def format_excel(df, title_text):
    """Applies Repair Kopitiam specific branding and layout to Excel."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Start data at Row 2 to leave room for Title
        df.to_excel(writer, index=False, sheet_name='Event Report', startrow=1)
        
        wb = writer.book
        ws = writer.sheets['Event Report']
        
        # Styles
        bold_font = Font(bold=True, size=12)
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Row 1: The Event Title
        ws.cell(row=1, column=1, value=title_text)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws.cell(row=1, column=1).font = bold_font
        
        # Row 2: Headers
        for cell in ws[2]:
            cell.font = bold_font
            cell.border = thin_border

        # Data Rows: Borders and Alignment
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border

        # Summary Section (10-row gap)
        start_summary = ws.max_row + 2
        ws.cell(row=start_summary, column=1, value="Walk-INs (Extra Rows below for manual entry)").font = bold_font
        
        # Add 10 empty rows with borders for Walk-ins
        for r in range(start_summary + 1, start_summary + 11):
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=r, column=c).border = thin_border
        
        # Final Statistics Footer
        stats_start = start_summary + 12
        stats_labels = [
            "Total approved registration for the event",
            "Walk-in Registrations",
            "No show",
            "Total attended",
            "Total items Fixed"
        ]
        for i, label in enumerate(stats_labels):
            cell = ws.cell(row=stats_start + i, column=1, value=label)
            cell.font = bold_font
            # Add a box for the number next to the label
            ws.cell(row=stats_start + i, column=2).border = thin_border

    return output.getvalue()

# --- 3. SIDEBAR CONTROLS ---
st.sidebar.image("https://repairkopitiam.sg/wp-content/uploads/2021/07/RK-Logo-1.png", width=150)
st.sidebar.title("Report Settings")

MONTHS = ["January", "February", "March", "April", "May", "June", 
          "July", "August", "September", "October", "November", "December"]

selected_month_name = st.sidebar.selectbox(
    "Select Event Month", 
    MONTHS,
    index=datetime.now().month - 1
)
month_int = MONTHS.index(selected_month_name) + 1
selected_year = st.sidebar.number_input("Event Year", min_value=2020, max_value=2030, value=datetime.now().year)

st.sidebar.divider()
st.sidebar.info("""
**How to use:**
1. Export CSV from RK Portal.
2. Upload file here.
3. Review preview.
4. Download Formatted Excel.
""")

# --- 4. MAIN INTERFACE ---
st.title("🛠️ Event Data Consolidator")
st.write("Transform raw portal exports into formatted National Repair Day reports instantly.")

uploaded_file = st.file_uploader("Upload Portal CSV/Excel Export", type=["csv", "xlsx"])

if uploaded_file:
    # Load data
    try:
        if uploaded_file.name.endswith('.csv'):
            raw_df = pd.read_csv(uploaded_file)
        else:
            raw_df = pd.read_excel(uploaded_file)
        
        # Standardize columns
        raw_df.columns = [c.strip() for c in raw_df.columns]
        
        # Filter Logic
        date_col = next((c for c in raw_df.columns if "date" in c.lower()), None)
        status_col = next((c for c in raw_df.columns if "status" in c.lower()), None)
        time_col = next((c for c in raw_df.columns if "time" in c.lower()), None)

        if date_col and status_col:
            raw_df[date_col] = pd.to_datetime(raw_df[date_col], errors='coerce')
            
            # Filter for Approved status and correct Date
            processed_df = raw_df[
                (raw_df[status_col].astype(str).str.strip().str.lower() == "approved") &
                (raw_df[date_col].dt.month == month_int) &
                (raw_df[date_col].dt.year == selected_year)
            ].copy()

            if not processed_df.empty:
                # Chronological Sorting
                if time_col:
                    processed_df['_sort_time'] = pd.to_datetime(processed_df[time_col].astype(str), errors='coerce').dt.time
                    processed_df = processed_df.sort_values(by='_sort_time').drop(columns=['_sort_time'])
                
                processed_df = processed_df.reset_index(drop=True)
                
                # Build the Output DataFrame
                final_df = pd.DataFrame()
                final_df['Comment'] = [""] * len(processed_df)
                final_df['Q.No'] = ""
                final_df['S.No'] = range(1, len(processed_df) + 1)
                final_df['User'] = processed_df['User'] if 'User' in processed_df.columns else "N/A"
                final_df['Phone'] = processed_df['Phone'] if 'Phone' in processed_df.columns else ""
                final_df['Time'] = processed_df[time_col] if time_col else ""
                final_df['Item 1'] = processed_df['Item 1'] if 'Item 1' in processed_df.columns else ""
                final_df['Item 1 Faults'] = "Not Working"
                final_df['Item 2'] = processed_df['Item 2'] if 'Item 2' in processed_df.columns else ""
                final_df['Item 2 Faults'] = processed_df['Item 2'].apply(lambda x: "Not Working" if pd.notnull(x) and x != "" else "")
                final_df['Total Items'] = ""
                final_df['Items Repaired'] = ""

                # Display Results
                st.subheader("📊 Data Preview")
                event_date_str = processed_df[date_col].iloc[0].strftime('%B %d, %Y')
                report_title = f"Repair Kopitiam@Coral Ris- National Repair Day {event_date_str}"
                
                st.success(f"Successfully processed {len(final_df)} approved records for {event_date_str}.")
                st.dataframe(final_df, use_container_width=True)

                # Excel Download
                excel_data = format_excel(final_df, report_title)
                
                st.download_button(
                    label="📥 Download Professional Excel Report",
                    data=excel_data,
                    file_name=f"RK_Report_{selected_month_name}_{selected_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning(f"No 'Approved' records found for {selected_month_name} {selected_year}.")
        else:
            st.error("Missing required columns: 'Event Date' or 'Status'. Check your export file.")

    except Exception as e:
        st.error(f"An error occurred: {e}")
        st.exception(e)

else:
    # Empty State
    st.info("👆 Upload your Portal CSV/Excel file to get started")
    
    # Sample data format guide
    with st.expander("📋 Expected CSV Format"):
        st.markdown("""
        Your CSV/Excel file should contain these columns:
        - **Event Date** - Date of the event (e.g., 2026-03-15)
        - **Status** - Registration status (should be "Approved")
        - **Time** - Time slot (e.g., 10:00 AM)
        - **User** - Participant name
        - **Phone** - Contact number
        - **Item 1** - First item to repair
        - **Item 2** - Second item (optional)
        """)
        
        sample_data = {
            'Event Date': ['2026-03-15', '2026-03-15'],
            'Status': ['Approved', 'Approved'],
            'Time': ['10:00 AM', '11:00 AM'],
            'User': ['John Doe', 'Jane Smith'],
            'Phone': ['91234567', '98765432'],
            'Item 1': ['Kettle', 'Fan'],
            'Item 2': ['', 'Lamp']
        }
        st.dataframe(pd.DataFrame(sample_data))